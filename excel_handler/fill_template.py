import os
import shutil
import openpyxl
import win32com.client as win32


TEMPLATE_PATH = os.path.join("excel_template", "bill_template.xlsx")


import os
import shutil
import openpyxl
import win32com.client as win32


TEMPLATE_PATH = os.path.join("excel_template", "bill_template.xlsx")


def fill_excel_template(data, output_path):
    """Fill the Excel bill template with the dynamic values"""

    # Copy blank template → new filled template
    shutil.copy(TEMPLATE_PATH, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # -------- STUDENT COPY ONLY --------

    ws["A6"]  = f"College/Institute Name : {data['college_name']}"
    ws["A8"]  = f"Ref. No. : {data['ref_no']}"
    ws["A9"]  = f"Date : {data['date']}"

    ws["A11"] = f"Student Name : {data['student_name']}"
    ws["A7"]  = f"For : {data['class']}"

    ws["B13"] = data["receipt1"]
    ws["B14"] = data["receipt2"]

    ws["B15"] = data["total_fees"]
    ws["B16"] = data["masoom_contribution"]
    ws["B17"] = data["student_contribution"]
    ws["B19"] = data["payable"]

    ws["A20"] = f"Rs. In Words : {data['rs_in_words']}"

    ws["A21"] = f"Donor Name : {data['donor_name']}"
    ws["A22"] = f"Cheque Issue On Name : {data['cheque_issue_name']}"

    ws["B23"] = f"A/C Holder Name : {data['account_holder']}"
    ws["B24"] = f"Bank Name : {data['bank_name']}"
    ws["B25"] = f"Account No : {data['account_number']}"
    ws["B26"] = f"IFSC Code : {data['ifsc']}"

    ws["A31"] = f"Cash/Cheque No : {data['cheque_number']}"

    wb.save(output_path)



def excel_to_pdf(input_path, output_path):
    """Convert filled Excel to a full-page A4 PDF with correct scaling."""
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(input_path))
    ws = wb.ActiveSheet

    try:
        # ---------------------
        # PAGE SETUP FOR A4 FULL PAGE
        # ---------------------
        ws.PageSetup.PaperSize = 9             # xlPaperA4
        ws.PageSetup.Orientation = 1           # 2 = Landscape, 1 = Portrait
        ws.PageSetup.Zoom = False              # Disable default zoom

        # Fit EVERYTHING into exactly 1 A4 page
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = 1

        # Thin margins for full coverage
        ws.PageSetup.LeftMargin   = excel.InchesToPoints(0.20)
        ws.PageSetup.RightMargin  = excel.InchesToPoints(0.20)
        ws.PageSetup.TopMargin    = excel.InchesToPoints(0.20)
        ws.PageSetup.BottomMargin = excel.InchesToPoints(0.20)

        # Remove header/footer space
        ws.PageSetup.HeaderMargin = excel.InchesToPoints(0.05)
        ws.PageSetup.FooterMargin = excel.InchesToPoints(0.05)

        # Set the print area EXACTLY (adjust if your template is tall/wide)
        ws.PageSetup.PrintArea = "$A$1:$B$33"

        # ---------------------
        # EXPORT
        # ---------------------
        wb.ExportAsFixedFormat(0, os.path.abspath(output_path))

    finally:
        wb.Close()
        excel.Quit()



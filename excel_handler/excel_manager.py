import os
import openpyxl
from datetime import datetime

RECORD_FILE = os.path.join("output", "records.xlsx")


def ensure_excel_file():
    """Create records.xlsx if not exists"""
    if not os.path.exists(RECORD_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Records"

        ws.append([
            "Date", "Reference No", "Student Name", "College Name", "Class",
            "Receipt 1", "Receipt 2", "Total Fees",
            "Masoom 75%", "Student 25%", "Payable", "Amount in Words",
            "Donor Name", "Cheque Issue Name", "Account Holder",
            "Bank Name", "Account Number", "IFSC", "Cheque Number",
            "Prepared By"
        ])

        wb.save(RECORD_FILE)


def load_donors_from_pdf():
    """Your old donor PDF function — leave empty if unused."""
    return []


def get_last_reference_number():
    """Fetch last reference number from records.xlsx"""

    ensure_excel_file()

    wb = openpyxl.load_workbook(RECORD_FILE)
    ws = wb.active

    if ws.max_row < 2:
        return None

    last_ref = ws.cell(ws.max_row, 2).value
    return last_ref


def next_ref_no():
    """
    Generate next reference number in format:
    M-0001/24-25/LT
    """

    last_ref = get_last_reference_number()
    year = datetime.now().strftime("%y")
    next_year = str(int(year) + 1)
    year_range = f"{year}-{next_year}"

    if not last_ref:
        return f"M-0001/{year_range}/LT"

    try:
        num_part = last_ref.split("-")[1].split("/")[0]
        num = int(num_part)
    except:
        num = 0

    num += 1
    return f"M-{num:04d}/{year_range}/LT"


def save_to_excel(data: dict):
    """Save the filled data into records.xlsx"""

    ensure_excel_file()

    wb = openpyxl.load_workbook(RECORD_FILE)
    ws = wb.active

    ws.append([
        data["date"],
        data["ref_no"],
        data["student_name"],
        data["college_name"],
        data["class"],
        data["receipt1"],
        data["receipt2"],
        data["total_fees"],
        data["masoom_contribution"],
        data["student_contribution"],
        data["payable"],
        data["rs_in_words"],
        data["donor_name"],
        data["cheque_issue_name"],
        data["account_holder"],
        data["bank_name"],
        data["account_number"],
        data["ifsc"],
        data["cheque_number"],
        data["prepared_by"]
    ])

    wb.save(RECORD_FILE)
